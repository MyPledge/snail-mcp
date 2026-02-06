import json
import sys
from pathlib import Path
from typing import Any, NamedTuple

import openpyxl

DEFAULT_OUT = "vbf"
DEFAULT_FMT = "j"


class _CellData(NamedTuple):
    row: int
    col: int
    value: str
    is_formula: bool
    bold: bool
    font_name: str
    font_size: Any
    font_color: str
    fill_color: str


def _font_color(cell) -> str:
    if not cell.font or not cell.font.color:
        return ""
    return getattr(cell.font.color, "rgb", None) or ""


def _fill_color(cell) -> str:
    if not cell.fill or not cell.fill.fgColor:
        return ""
    return getattr(cell.fill.fgColor, "rgb", None) or ""


def _cell_value_str(cell) -> str:
    if cell.value is None:
        return ""
    return cell.value if isinstance(cell.value, str) else str(cell.value)


def _esc(s: str) -> str:
    if not s:
        return s
    return s.replace("\\", "\\\\").replace("\t", "\\t").replace("\r", "\\r").replace("\n", "\\n")


def _parse_out(out: str) -> dict[str, bool]:
    o = (out or DEFAULT_OUT).lower()
    return {
        "v": "v" in o,
        "b": "b" in o,
        "n": "n" in o,
        "z": "z" in o,
        "c": "c" in o,
        "g": "g" in o,
        "f": "f" in o,
    }


def _is_blank(val: Any) -> bool:
    return val is None or (isinstance(val, str) and not val.strip())


def _sheet_range(ws, row_start: int, row_end: int, col_start: int, col_end: int) -> tuple[int, int, int, int]:
    r1 = row_start if row_start > 0 else (ws.min_row or 1)
    r2 = row_end if row_end > 0 else (ws.max_row or 1)
    c1 = col_start if col_start > 0 else (ws.min_column or 1)
    c2 = col_end if col_end > 0 else (ws.max_column or 1)
    return r1, r2, c1, c2


def _collect_cells(ws, r1: int, r2: int, c1: int, c2: int, sparse: bool) -> list[_CellData]:
    cells: list[_CellData] = []
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if sparse and _is_blank(val):
                continue
            cells.append(_CellData(
                row=row,
                col=col,
                value=_cell_value_str(cell),
                is_formula=isinstance(val, str) and val.startswith("="),
                bold=cell.font.bold if cell.font else False,
                font_name=cell.font.name if cell.font else "",
                font_size=cell.font.size if cell.font else None,
                font_color=_font_color(cell),
                fill_color=_fill_color(cell),
            ))
    return cells


def _cell_to_json(cell: _CellData, fields: dict[str, bool]) -> dict[str, Any]:
    obj: dict[str, Any] = {"row": cell.row, "col": cell.col}
    if fields["v"]:
        obj["value"] = cell.value
    if fields["f"] and cell.is_formula:
        obj["formula"] = True
    if fields["b"]:
        obj["bold"] = cell.bold
    if fields["n"]:
        obj["font_name"] = cell.font_name or ""
    if fields["z"]:
        obj["font_size"] = cell.font_size
    if fields["c"]:
        obj["font_color"] = cell.font_color
    if fields["g"]:
        obj["fill_color"] = cell.fill_color
    return obj


def _cell_to_text_parts(cell: _CellData, fields: dict[str, bool]) -> list[str]:
    parts = ["c", str(cell.row), str(cell.col)]
    if fields["v"]:
        parts.append(_esc(cell.value))
    if fields["f"]:
        parts.append("1" if cell.is_formula else "0")
    if fields["b"]:
        parts.append("1" if cell.bold else "0")
    if fields["n"]:
        parts.append(_esc(cell.font_name or ""))
    if fields["z"]:
        parts.append(str(cell.font_size) if cell.font_size else "")
    if fields["c"]:
        parts.append(_esc(cell.font_color))
    if fields["g"]:
        parts.append(_esc(cell.fill_color))
    return parts


def read_excel(
    path: str,
    out: str = DEFAULT_OUT,
    fmt: str = DEFAULT_FMT,
    sheets: str = "",
    row_start: int = 0,
    row_end: int = 0,
    col_start: int = 0,
    col_end: int = 0,
    sparse: bool = True,
) -> str:
    fields = _parse_out(out)
    fmt_char = (fmt or DEFAULT_FMT).lower()[:1]
    sheet_names = [s.strip() for s in sheets.split(",") if s.strip()]

    wb = openpyxl.load_workbook(str(path), data_only=False)
    try:
        sheet_names = sheet_names or wb.sheetnames
        cells_by_sheet: dict[str, list[_CellData]] = {}
        for name in sheet_names:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            r1, r2, c1, c2 = _sheet_range(ws, row_start, row_end, col_start, col_end)
            if r2 < r1 or c2 < c1:
                continue
            cells_by_sheet[name] = _collect_cells(ws, r1, r2, c1, c2, sparse)
    finally:
        wb.close()

    if fmt_char == "j":
        result = {}
        for name in sheet_names:
            if name in cells_by_sheet:
                result[name] = {"cells": [_cell_to_json(c, fields) for c in cells_by_sheet[name]]}
        return json.dumps(result, ensure_ascii=False)

    lines: list[str] = []
    for name in sheet_names:
        if name not in cells_by_sheet:
            continue
        lines.append(">" + _esc(name))
        for cell in cells_by_sheet[name]:
            lines.append("\t".join(_cell_to_text_parts(cell, fields)))
    return "\n".join(lines)


def _main() -> None:
    path = (sys.argv[1] if len(sys.argv) > 1 else "").strip()
    out = (sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT).strip()
    fmt = (sys.argv[3] if len(sys.argv) > 3 else DEFAULT_FMT).strip()
    if not path or not Path(path).exists():
        print("usage: read_excel <path> [out=vbf] [fmt=t|j]\nout: v=value b=bold n=font_name z=font_size c=font_color g=fill f=formula. range: 0=used. sparse=0|1.", file=sys.stderr)
        sys.exit(1)
    print(read_excel(path, out=out, fmt=fmt))


if __name__ == "__main__":
    _main()
